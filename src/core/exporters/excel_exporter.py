"""
Excel Exporter - Esportazione risultati in formato Excel
Calcolatore Cerchiature NTC 2018
Arch. Michelangelo Bartolotta
"""

import os
from datetime import datetime

# Prova openpyxl, altrimenti usa csv come fallback
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import csv
import json


class ExcelExporter:
    """Esportatore risultati in formato Excel/CSV"""

    def __init__(self):
        self.workbook = None
        self.styles = {}
        if OPENPYXL_AVAILABLE:
            self._setup_styles()

    def _setup_styles(self):
        """Configura stili per Excel"""
        # Font
        self.styles['title_font'] = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.styles['header_font'] = Font(name='Arial', size=11, bold=True)
        self.styles['normal_font'] = Font(name='Arial', size=10)
        self.styles['small_font'] = Font(name='Arial', size=9, italic=True)

        # Allineamenti
        self.styles['center'] = Alignment(horizontal='center', vertical='center')
        self.styles['left'] = Alignment(horizontal='left', vertical='center')
        self.styles['right'] = Alignment(horizontal='right', vertical='center')

        # Bordi
        thin_border = Side(style='thin', color='000000')
        self.styles['border'] = Border(
            left=thin_border, right=thin_border,
            top=thin_border, bottom=thin_border
        )

        # Riempimenti
        self.styles['header_fill'] = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        self.styles['success_fill'] = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
        self.styles['error_fill'] = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
        self.styles['alt_fill'] = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')

    def export(self, filepath, project_data, results):
        """
        Esporta i dati del progetto e i risultati in Excel

        Args:
            filepath: Percorso file di destinazione
            project_data: Dizionario con dati progetto
            results: Dizionario con risultati calcolo
        """
        if OPENPYXL_AVAILABLE:
            return self._export_xlsx(filepath, project_data, results)
        else:
            # Fallback a CSV
            return self._export_csv(filepath, project_data, results)

    def _export_xlsx(self, filepath, project_data, results):
        """Esporta in formato Excel XLSX"""
        self.workbook = Workbook()

        # Rimuovi foglio predefinito
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        # Crea fogli
        self._create_summary_sheet(project_data, results)
        self._create_geometry_sheet(project_data)
        self._create_openings_sheet(project_data)
        self._create_results_sheet(results)
        self._create_verification_sheet(results)

        # Salva
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'

        self.workbook.save(filepath)
        return filepath

    def _create_summary_sheet(self, project_data, results):
        """Crea foglio riepilogo"""
        ws = self.workbook.create_sheet("Riepilogo")

        # Titolo
        ws.merge_cells('A1:F1')
        ws['A1'] = "CALCOLO CERCHIATURE NTC 2018 - RIEPILOGO"
        ws['A1'].font = self.styles['title_font']
        ws['A1'].fill = self.styles['header_fill']
        ws['A1'].alignment = self.styles['center']

        # Info progetto
        row = 3
        ws[f'A{row}'] = "Data calcolo:"
        ws[f'B{row}'] = datetime.now().strftime("%d/%m/%Y %H:%M")

        row += 1
        ws[f'A{row}'] = "Progetto:"
        ws[f'B{row}'] = project_data.get('project_info', {}).get('name', 'Non specificato')

        row += 1
        ws[f'A{row}'] = "Committente:"
        ws[f'B{row}'] = project_data.get('project_info', {}).get('client', '-')

        # Geometria muro
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "GEOMETRIA MURO"
        ws[f'A{row}'].font = self.styles['header_font']
        ws[f'A{row}'].fill = self.styles['alt_fill']

        wall = project_data.get('wall', {})
        row += 1
        ws[f'A{row}'] = "Lunghezza [cm]:"
        ws[f'B{row}'] = wall.get('length', 0)
        ws[f'C{row}'] = "Altezza [cm]:"
        ws[f'D{row}'] = wall.get('height', 0)
        ws[f'E{row}'] = "Spessore [cm]:"
        ws[f'F{row}'] = wall.get('thickness', 0)

        # Risultato verifica
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        verification = results.get('verification', {})
        if verification.get('is_local'):
            ws[f'A{row}'] = "VERIFICA: INTERVENTO LOCALE"
            ws[f'A{row}'].fill = self.styles['success_fill']
        else:
            ws[f'A{row}'] = "VERIFICA: NON LOCALE"
            ws[f'A{row}'].fill = self.styles['error_fill']
        ws[f'A{row}'].font = self.styles['title_font']
        ws[f'A{row}'].alignment = self.styles['center']

        # Variazioni
        row += 2
        ws[f'A{row}'] = "Variazione rigidezza ΔK:"
        ws[f'B{row}'] = f"{verification.get('stiffness_variation', 0):.2f}%"
        ws[f'C{row}'] = "Limite:"
        ws[f'D{row}'] = "±15%"

        row += 1
        ws[f'A{row}'] = "Variazione resistenza ΔV:"
        ws[f'B{row}'] = f"{verification.get('resistance_variation', 0):.2f}%"
        ws[f'C{row}'] = "Limite:"
        ws[f'D{row}'] = "±15%"

        # Larghezza colonne
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_geometry_sheet(self, project_data):
        """Crea foglio geometria"""
        ws = self.workbook.create_sheet("Geometria")

        # Header
        headers = ['Parametro', 'Valore', 'Unità', 'Note']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.styles['header_font']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center']
            cell.border = self.styles['border']

        # Dati muro
        wall = project_data.get('wall', {})
        data = [
            ('Lunghezza muro', wall.get('length', 0), 'cm', ''),
            ('Altezza muro', wall.get('height', 0), 'cm', ''),
            ('Spessore muro', wall.get('thickness', 0), 'cm', ''),
            ('Area lorda', wall.get('length', 0) * wall.get('height', 0) / 10000, 'm²', 'L × H'),
        ]

        # Dati muratura
        masonry = project_data.get('masonry', {})
        data.extend([
            ('', '', '', ''),
            ('Tipo muratura', masonry.get('type', '-'), '', ''),
            ('fcm', masonry.get('fcm', 0), 'MPa', 'Resistenza media compressione'),
            ('tau0', masonry.get('tau0', 0), 'MPa', 'Resistenza media taglio'),
            ('E', masonry.get('E', 0), 'MPa', 'Modulo elastico'),
            ('G', masonry.get('G', 0), 'MPa', 'Modulo taglio'),
            ('Livello conoscenza', masonry.get('knowledge_level', '-'), '', ''),
            ('FC', masonry.get('FC', 1.35), '', 'Fattore confidenza'),
        ])

        for row, (param, value, unit, note) in enumerate(data, 2):
            ws.cell(row=row, column=1, value=param).border = self.styles['border']
            ws.cell(row=row, column=2, value=value).border = self.styles['border']
            ws.cell(row=row, column=3, value=unit).border = self.styles['border']
            ws.cell(row=row, column=4, value=note).border = self.styles['border']

        # Larghezza colonne
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30

    def _create_openings_sheet(self, project_data):
        """Crea foglio aperture"""
        ws = self.workbook.create_sheet("Aperture")

        # Header
        headers = ['N.', 'Tipo', 'X [cm]', 'Y [cm]', 'Largh. [cm]', 'Alt. [cm]', 'Area [m²]', 'Rinforzo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.styles['header_font']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center']
            cell.border = self.styles['border']

        # Dati aperture
        openings = project_data.get('openings', [])
        if not openings:
            # Prova in openings_module
            openings_data = project_data.get('openings_module', {})
            openings = openings_data.get('openings', [])

        for row, opening in enumerate(openings, 2):
            ws.cell(row=row, column=1, value=row-1).border = self.styles['border']
            ws.cell(row=row, column=2, value=opening.get('type', 'Rettangolare')).border = self.styles['border']
            ws.cell(row=row, column=3, value=opening.get('x', 0)).border = self.styles['border']
            ws.cell(row=row, column=4, value=opening.get('y', 0)).border = self.styles['border']
            ws.cell(row=row, column=5, value=opening.get('width', 0)).border = self.styles['border']
            ws.cell(row=row, column=6, value=opening.get('height', 0)).border = self.styles['border']

            area = opening.get('width', 0) * opening.get('height', 0) / 10000
            ws.cell(row=row, column=7, value=round(area, 3)).border = self.styles['border']

            rinforzo = opening.get('rinforzo', {})
            if rinforzo:
                rinf_text = f"{rinforzo.get('profilo', '-')} ({rinforzo.get('tipo_acciaio', 'S275')})"
            else:
                rinf_text = "Nessuno"
            ws.cell(row=row, column=8, value=rinf_text).border = self.styles['border']

        # Se nessuna apertura
        if not openings:
            ws.cell(row=2, column=1, value="Nessuna apertura definita")
            ws.merge_cells('A2:H2')

        # Larghezza colonne
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 14

    def _create_results_sheet(self, results):
        """Crea foglio risultati calcolo"""
        ws = self.workbook.create_sheet("Risultati")

        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = "RISULTATI CALCOLO"
        ws['A1'].font = self.styles['title_font']
        ws['A1'].fill = self.styles['header_fill']
        ws['A1'].alignment = self.styles['center']

        # Stato iniziale
        row = 3
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "STATO INIZIALE (ante-operam)"
        ws[f'A{row}'].font = self.styles['header_font']
        ws[f'A{row}'].fill = self.styles['alt_fill']

        initial = results.get('initial', {})
        row += 1
        ws[f'A{row}'] = "Rigidezza K₀:"
        ws[f'B{row}'] = f"{initial.get('K', 0):.2f}"
        ws[f'C{row}'] = "kN/m"

        row += 1
        ws[f'A{row}'] = "Resistenza V_t1:"
        ws[f'B{row}'] = f"{initial.get('V_t1', 0):.2f}"
        ws[f'C{row}'] = "kN"

        row += 1
        ws[f'A{row}'] = "Resistenza V_t2:"
        ws[f'B{row}'] = f"{initial.get('V_t2', 0):.2f}"
        ws[f'C{row}'] = "kN"

        row += 1
        ws[f'A{row}'] = "Resistenza V_t3:"
        ws[f'B{row}'] = f"{initial.get('V_t3', 0):.2f}"
        ws[f'C{row}'] = "kN"

        # Stato modificato
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "STATO MODIFICATO (post-operam)"
        ws[f'A{row}'].font = self.styles['header_font']
        ws[f'A{row}'].fill = self.styles['alt_fill']

        modified = results.get('modified', {})
        row += 1
        ws[f'A{row}'] = "Rigidezza K_mod:"
        ws[f'B{row}'] = f"{modified.get('K', 0):.2f}"
        ws[f'C{row}'] = "kN/m"

        row += 1
        ws[f'A{row}'] = "Resistenza V_t1:"
        ws[f'B{row}'] = f"{modified.get('V_t1', 0):.2f}"
        ws[f'C{row}'] = "kN"

        row += 1
        ws[f'A{row}'] = "Resistenza V_t2:"
        ws[f'B{row}'] = f"{modified.get('V_t2', 0):.2f}"
        ws[f'C{row}'] = "kN"

        row += 1
        ws[f'A{row}'] = "Resistenza V_t3:"
        ws[f'B{row}'] = f"{modified.get('V_t3', 0):.2f}"
        ws[f'C{row}'] = "kN"

        # Contributo cerchiature
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "CONTRIBUTO CERCHIATURE"
        ws[f'A{row}'].font = self.styles['header_font']
        ws[f'A{row}'].fill = self.styles['alt_fill']

        cerchiature = results.get('cerchiature', {})
        row += 1
        ws[f'A{row}'] = "K cerchiature:"
        ws[f'B{row}'] = f"{cerchiature.get('K_total', 0):.2f}"
        ws[f'C{row}'] = "kN/m"

        row += 1
        ws[f'A{row}'] = "V cerchiature:"
        ws[f'B{row}'] = f"{cerchiature.get('V_total', 0):.2f}"
        ws[f'C{row}'] = "kN"

        # Larghezza colonne
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20

    def _create_verification_sheet(self, results):
        """Crea foglio verifica"""
        ws = self.workbook.create_sheet("Verifica")

        verification = results.get('verification', {})

        # Titolo
        ws.merge_cells('A1:D1')
        ws['A1'] = "VERIFICA INTERVENTO LOCALE (C8.4.3 NTC 2018)"
        ws['A1'].font = self.styles['title_font']
        ws['A1'].fill = self.styles['header_fill']
        ws['A1'].alignment = self.styles['center']

        # Esito
        row = 3
        ws.merge_cells(f'A{row}:D{row}')
        if verification.get('is_local'):
            ws[f'A{row}'] = "ESITO: INTERVENTO LOCALE VERIFICATO"
            ws[f'A{row}'].fill = self.styles['success_fill']
        else:
            ws[f'A{row}'] = "ESITO: INTERVENTO NON LOCALE"
            ws[f'A{row}'].fill = self.styles['error_fill']
        ws[f'A{row}'].font = self.styles['title_font']
        ws[f'A{row}'].alignment = self.styles['center']

        # Dettagli
        row = 5
        headers = ['Parametro', 'Valore', 'Limite', 'Esito']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles['header_font']
            cell.fill = self.styles['alt_fill']
            cell.border = self.styles['border']

        # Rigidezza
        row += 1
        ws.cell(row=row, column=1, value="Variazione rigidezza ΔK").border = self.styles['border']
        ws.cell(row=row, column=2, value=f"{verification.get('stiffness_variation', 0):.2f}%").border = self.styles['border']
        ws.cell(row=row, column=3, value="±15%").border = self.styles['border']
        stiff_ok = verification.get('stiffness_ok', False)
        cell = ws.cell(row=row, column=4, value="OK" if stiff_ok else "NON OK")
        cell.border = self.styles['border']
        cell.fill = self.styles['success_fill'] if stiff_ok else self.styles['error_fill']
        cell.font = Font(color='FFFFFF', bold=True)

        # Resistenza
        row += 1
        ws.cell(row=row, column=1, value="Variazione resistenza ΔV").border = self.styles['border']
        ws.cell(row=row, column=2, value=f"{verification.get('resistance_variation', 0):.2f}%").border = self.styles['border']
        ws.cell(row=row, column=3, value="±15%").border = self.styles['border']
        res_ok = verification.get('resistance_ok', False)
        cell = ws.cell(row=row, column=4, value="OK" if res_ok else "NON OK")
        cell.border = self.styles['border']
        cell.fill = self.styles['success_fill'] if res_ok else self.styles['error_fill']
        cell.font = Font(color='FFFFFF', bold=True)

        # Note normativa
        row += 3
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "Riferimento normativo:"
        ws[f'A{row}'].font = self.styles['header_font']

        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = ("C8.4.3 NTC 2018: L'intervento locale non deve modificare "
                        "significativamente rigidezza e resistenza dell'elemento (±15%).")
        ws[f'A{row}'].font = self.styles['small_font']

        # Larghezza colonne
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12

    def _export_csv(self, filepath, project_data, results):
        """Fallback: esporta in formato CSV"""
        if not filepath.endswith('.csv'):
            filepath = filepath.replace('.xlsx', '.csv')

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')

            # Header
            writer.writerow(['CALCOLO CERCHIATURE NTC 2018'])
            writer.writerow(['Data', datetime.now().strftime("%d/%m/%Y %H:%M")])
            writer.writerow([])

            # Geometria
            wall = project_data.get('wall', {})
            writer.writerow(['GEOMETRIA MURO'])
            writer.writerow(['Lunghezza [cm]', wall.get('length', 0)])
            writer.writerow(['Altezza [cm]', wall.get('height', 0)])
            writer.writerow(['Spessore [cm]', wall.get('thickness', 0)])
            writer.writerow([])

            # Risultati
            writer.writerow(['RISULTATI'])
            initial = results.get('initial', {})
            writer.writerow(['K iniziale [kN/m]', initial.get('K', 0)])
            writer.writerow(['V_t1 iniziale [kN]', initial.get('V_t1', 0)])

            modified = results.get('modified', {})
            writer.writerow(['K modificato [kN/m]', modified.get('K', 0)])
            writer.writerow(['V_t1 modificato [kN]', modified.get('V_t1', 0)])
            writer.writerow([])

            # Verifica
            verification = results.get('verification', {})
            writer.writerow(['VERIFICA'])
            writer.writerow(['Variazione K [%]', verification.get('stiffness_variation', 0)])
            writer.writerow(['Variazione V [%]', verification.get('resistance_variation', 0)])
            writer.writerow(['Esito', 'LOCALE' if verification.get('is_local') else 'NON LOCALE'])

        return filepath


def export_to_excel(filepath, project_data, results):
    """Funzione helper per esportazione rapida"""
    exporter = ExcelExporter()
    return exporter.export(filepath, project_data, results)
